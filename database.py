import sqlite3
import os
from datetime import datetime

# Nama file database
DB_FILE = 'catering_finance.db'

def init_database():
    """Inisialisasi database dan membuat tabel jika belum ada"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Membuat tabel transaksi keuangan
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            day TEXT NOT NULL,
            description TEXT NOT NULL,
            income REAL DEFAULT 0,
            expense REAL DEFAULT 0,
            balance REAL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Membuat tabel ringkasan keuangan
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS finance_summary (
            id INTEGER PRIMARY KEY,
            total_income REAL DEFAULT 0,
            total_expense REAL DEFAULT 0,
            current_balance REAL DEFAULT 0,
            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Memastikan ada satu baris di tabel summary
    cursor.execute('SELECT COUNT(*) FROM finance_summary')
    if cursor.fetchone()[0] == 0:
        cursor.execute('''
            INSERT INTO finance_summary (id, total_income, total_expense, current_balance)
            VALUES (1, 0, 0, 0)
        ''')
    
    conn.commit()
    conn.close()
    print("Database initialized successfully!")

def add_income(date, day, description, amount):
    """Menambahkan pemasukan ke database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Mendapatkan saldo terakhir
    cursor.execute('SELECT current_balance FROM finance_summary WHERE id = 1')
    result = cursor.fetchone()
    previous_balance = result[0] if result else 0
    
    # Menghitung saldo baru
    new_balance = previous_balance + amount
    
    # Menyimpan transaksi
    cursor.execute('''
        INSERT INTO transactions (date, day, description, income, balance)
        VALUES (?, ?, ?, ?, ?)
    ''', (date, day, description, amount, new_balance))
    
    # Memperbarui ringkasan keuangan
    cursor.execute('''
        UPDATE finance_summary 
        SET total_income = total_income + ?, 
            current_balance = ?,
            last_updated = CURRENT_TIMESTAMP
        WHERE id = 1
    ''', (amount, new_balance))
    
    conn.commit()
    conn.close()
    return new_balance

def add_expense(date, day, description, amount):
    """Menambahkan pengeluaran ke database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Mendapatkan saldo terakhir
    cursor.execute('SELECT current_balance FROM finance_summary WHERE id = 1')
    result = cursor.fetchone()
    previous_balance = result[0] if result else 0
    
    # Menghitung saldo baru
    new_balance = previous_balance - amount
    
    # Menyimpan transaksi
    cursor.execute('''
        INSERT INTO transactions (date, day, description, expense, balance)
        VALUES (?, ?, ?, ?, ?)
    ''', (date, day, description, amount, new_balance))
    
    # Memperbarui ringkasan keuangan
    cursor.execute('''
        UPDATE finance_summary 
        SET total_expense = total_expense + ?, 
            current_balance = ?,
            last_updated = CURRENT_TIMESTAMP
        WHERE id = 1
    ''', (amount, new_balance))
    
    conn.commit()
    conn.close()
    return new_balance

def get_all_transactions():
    """Mendapatkan semua transaksi dari database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT date, day, description, income, expense, balance
        FROM transactions
        ORDER BY date, created_at
    ''')
    
    transactions = cursor.fetchall()
    conn.close()
    
    # Mengonversi ke list of dictionaries
    result = []
    for transaction in transactions:
        result.append({
            'date': transaction[0],
            'day': transaction[1],
            'description': transaction[2],
            'income': transaction[3],
            'expense': transaction[4],
            'balance': transaction[5]
        })
    
    return result

def get_finance_summary():
    """Mendapatkan ringkasan keuangan"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT total_income, total_expense, current_balance
        FROM finance_summary
        WHERE id = 1
    ''')
    
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return {
            'total_income': result[0],
            'total_expense': result[1],
            'current_balance': result[2]
        }
    else:
        return {
            'total_income': 0,
            'total_expense': 0,
            'current_balance': 0
        }

def get_day_name(date_str):
    """Mendapatkan nama hari dari tanggal (format: DD/MM/YYYY)"""
    days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    try:
        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
        return days[date_obj.weekday()]
    except ValueError:
        return "Tidak valid"

def migrate_from_excel():
    """Migrasi data dari file Excel yang ada (jika ada)"""
    import openpyxl
    
    excel_files = [
        'catering_finance_console.xlsx',
        'catering_finance_web.xlsx',
        'catering_finance_pwa.xlsx'
    ]
    
    for excel_file in excel_files:
        if os.path.exists(excel_file):
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook["Keuangan_Katering"] if "Keuangan_Katering" in workbook.sheetnames else workbook.active
                
                # Migrasi transaksi dari Excel ke database
                for row in range(2, worksheet.max_row + 1):
                    date = worksheet.cell(row=row, column=1).value or ""
                    day = worksheet.cell(row=row, column=2).value or ""
                    description = worksheet.cell(row=row, column=3).value or ""
                    income = worksheet.cell(row=row, column=4).value or 0
                    expense = worksheet.cell(row=row, column=5).value or 0
                    balance = worksheet.cell(row=row, column=6).value or 0
                    
                    # Cek apakah transaksi sudah ada
                    conn = sqlite3.connect(DB_FILE)
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT COUNT(*) FROM transactions 
                        WHERE date = ? AND description = ? AND income = ? AND expense = ?
                    ''', (date, description, income, expense))
                    
                    if cursor.fetchone()[0] == 0:
                        # Tambahkan transaksi jika belum ada
                        cursor.execute('''
                            INSERT INTO transactions (date, day, description, income, expense, balance)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (date, day, description, income, expense, balance))
                        conn.commit()
                    conn.close()
                
                print(f"Data migrated from {excel_file}")
            except Exception as e:
                print(f"Error migrating data from {excel_file}: {e}")

# Inisialisasi database saat modul diimpor
init_database()
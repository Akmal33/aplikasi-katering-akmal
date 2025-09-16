from flask import Flask, render_template, request, jsonify
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)

# Nama file Excel
EXCEL_FILE = 'catering_finance_pwa.xlsx'

def initialize_workbook():
    """Inisialisasi workbook dan worksheet"""
    if os.path.exists(EXCEL_FILE):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        if "Keuangan_Katering" in workbook.sheetnames:
            worksheet = workbook["Keuangan_Katering"]
        else:
            worksheet = workbook.create_sheet("Keuangan_Katering")
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Keuangan_Katering"
        create_headers(worksheet)
    
    # Simpan workbook
    workbook.save(EXCEL_FILE)
    return workbook, worksheet

def create_headers(worksheet):
    """Membuat header untuk worksheet"""
    headers = ["Tanggal", "Hari", "Deskripsi", "Pemasukan (Rp)", "Pengeluaran (Rp)", "Saldo (Rp)"]
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)

def get_day_name(date_str):
    """Mendapatkan nama hari dari tanggal"""
    days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        return days[date_obj.weekday()]
    except ValueError:
        return "Tidak valid"

def add_income(date, description, amount):
    """Menambahkan pemasukan"""
    workbook, worksheet = initialize_workbook()
    day = get_day_name(date)
    
    # Format tanggal untuk tampilan
    display_date = datetime.strptime(date, "%Y-%m-%d").strftime("%d/%m/%Y")
    
    next_row = worksheet.max_row + 1
    worksheet.cell(row=next_row, column=1, value=display_date)
    worksheet.cell(row=next_row, column=2, value=day)
    worksheet.cell(row=next_row, column=3, value=description)
    worksheet.cell(row=next_row, column=4, value=amount)
    worksheet.cell(row=next_row, column=5, value=0)
    
    # Hitung saldo baru
    if next_row == 2:  # Baris pertama dengan data
        previous_balance = 0
    else:
        previous_balance = worksheet.cell(row=next_row-1, column=6).value or 0
    
    current_balance = previous_balance + amount
    worksheet.cell(row=next_row, column=6, value=current_balance)
    
    workbook.save(EXCEL_FILE)
    return current_balance

def add_expense(date, description, amount):
    """Menambahkan pengeluaran"""
    workbook, worksheet = initialize_workbook()
    day = get_day_name(date)
    
    # Format tanggal untuk tampilan
    display_date = datetime.strptime(date, "%Y-%m-%d").strftime("%d/%m/%Y")
    
    next_row = worksheet.max_row + 1
    worksheet.cell(row=next_row, column=1, value=display_date)
    worksheet.cell(row=next_row, column=2, value=day)
    worksheet.cell(row=next_row, column=3, value=description)
    worksheet.cell(row=next_row, column=4, value=0)
    worksheet.cell(row=next_row, column=5, value=amount)
    
    # Hitung saldo baru
    if next_row == 2:  # Baris pertama dengan data
        previous_balance = 0
    else:
        previous_balance = worksheet.cell(row=next_row-1, column=6).value or 0
    
    current_balance = previous_balance - amount
    worksheet.cell(row=next_row, column=6, value=current_balance)
    
    workbook.save(EXCEL_FILE)
    return current_balance

def get_all_transactions():
    """Mendapatkan semua transaksi"""
    workbook, worksheet = initialize_workbook()
    transactions = []
    
    # Iterasi semua baris dengan data
    for row in range(2, worksheet.max_row + 1):
        transaction = {
            'date': worksheet.cell(row=row, column=1).value or "",
            'day': worksheet.cell(row=row, column=2).value or "",
            'description': worksheet.cell(row=row, column=3).value or "",
            'income': worksheet.cell(row=row, column=4).value or 0,
            'expense': worksheet.cell(row=row, column=5).value or 0,
            'balance': worksheet.cell(row=row, column=6).value or 0
        }
        transactions.append(transaction)
    
    return transactions

def calculate_totals():
    """Menghitung total pemasukan, pengeluaran, dan saldo"""
    workbook, worksheet = initialize_workbook()
    total_income = 0
    total_expense = 0
    current_balance = 0
    
    # Iterasi semua baris dengan data
    for row in range(2, worksheet.max_row + 1):
        income = worksheet.cell(row=row, column=4).value or 0
        expense = worksheet.cell(row=row, column=5).value or 0
        balance = worksheet.cell(row=row, column=6).value or 0
        
        total_income += income
        total_expense += expense
        current_balance = balance  # Saldo terakhir
    
    return total_income, total_expense, current_balance

@app.route('/')
def index():
    """Halaman utama"""
    transactions = get_all_transactions()
    total_income, total_expense, current_balance = calculate_totals()
    
    return render_template('index.html', 
                         transactions=transactions,
                         total_income=total_income,
                         total_expense=total_expense,
                         current_balance=current_balance)

@app.route('/add_income', methods=['POST'])
def add_income_route():
    """Route untuk menambahkan pemasukan"""
    try:
        data = request.get_json()
        date = data['date']
        description = data['description']
        amount = float(data['amount'])
        
        balance = add_income(date, description, amount)
        
        return jsonify({
            'status': 'success',
            'message': f'Pemasukan berhasil ditambahkan! Saldo terbaru: Rp {balance:,.0f}',
            'balance': balance
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Terjadi kesalahan: {str(e)}'
        }), 400

@app.route('/add_expense', methods=['POST'])
def add_expense_route():
    """Route untuk menambahkan pengeluaran"""
    try:
        data = request.get_json()
        date = data['date']
        description = data['description']
        amount = float(data['amount'])
        
        balance = add_expense(date, description, amount)
        
        return jsonify({
            'status': 'success',
            'message': f'Pengeluaran berhasil ditambahkan! Saldo terbaru: Rp {balance:,.0f}',
            'balance': balance
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Terjadi kesalahan: {str(e)}'
        }), 400

# Route untuk file statis PWA
@app.route('/manifest.json')
def manifest():
    return app.send_static_file('manifest.json')

@app.route('/service-worker.js')
def service_worker():
    return app.send_static_file('service-worker.js')

if __name__ == '__main__':
    # Inisialisasi workbook saat aplikasi dimulai
    initialize_workbook()
    app.run(debug=True, host='0.0.0.0', port=5000)
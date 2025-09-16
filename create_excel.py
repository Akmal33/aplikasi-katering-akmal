import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Kartu Stok"

# Add headers
headers = ["No.", "Tanggal", "Keterangan", "Masuk", "Keluar", "Sisa"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Add sample data based on the XML structure
sample_data = [
    [1, "01/09/2025", "Pembelian awal", 100, 0, 100],
    [2, "05/09/2025", "Barang terjual", 0, 20, 80],
    [3, "10/09/2025", "Pembelian tambahan", 50, 0, 130],
    [4, "15/09/2025", "Barang terjual", 0, 30, 100]
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Save the workbook
wb.save("kartu_stok_proper.xlsx")
print("New Excel file created successfully!")
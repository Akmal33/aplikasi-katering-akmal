from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
import openpyxl
import os
from datetime import datetime

class FinanceEntry(BoxLayout):
    def __init__(self, date, day, description, income, expense, balance, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'horizontal'
        self.size_hint_y = None
        self.height = '40dp'
        
        self.add_widget(Label(text=date, size_hint_x=0.15))
        self.add_widget(Label(text=day, size_hint_x=0.15))
        self.add_widget(Label(text=description, size_hint_x=0.3))
        self.add_widget(Label(text=f"Rp {income:,.0f}", size_hint_x=0.15))
        self.add_widget(Label(text=f"Rp {expense:,.0f}", size_hint_x=0.15))
        self.add_widget(Label(text=f"Rp {balance:,.0f}", size_hint_x=0.15))

class CateringFinanceApp(App):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.excel_file = "catering_finance_mobile.xlsx"
        self.initialize_workbook()
    
    def initialize_workbook(self):
        """Inisialisasi workbook dan worksheet"""
        if os.path.exists(self.excel_file):
            self.workbook = openpyxl.load_workbook(self.excel_file)
            if "Keuangan_Katering" in self.workbook.sheetnames:
                self.worksheet = self.workbook["Keuangan_Katering"]
            else:
                self.worksheet = self.workbook.create_sheet("Keuangan_Katering")
        else:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Keuangan_Katering"
            self.create_headers()
    
    def create_headers(self):
        """Membuat header untuk worksheet"""
        headers = ["Tanggal", "Hari", "Deskripsi", "Pemasukan (Rp)", "Pengeluaran (Rp)", "Saldo (Rp)"]
        for col, header in enumerate(headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)
    
    def get_day_name(self, date_str):
        """Mendapatkan nama hari dari tanggal"""
        days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        try:
            date_obj = datetime.strptime(date_str, "%d/%m/%Y")
            return days[date_obj.weekday()]
        except ValueError:
            return "Tidak valid"
    
    def calculate_and_update_balance(self, row):
        """Menghitung dan memperbarui saldo"""
        current_income = self.worksheet.cell(row=row, column=4).value or 0
        current_expense = self.worksheet.cell(row=row, column=5).value or 0
        
        if row == 2:  # Baris pertama dengan data
            previous_balance = 0
        else:
            previous_balance = self.worksheet.cell(row=row-1, column=6).value or 0
        
        current_balance = previous_balance + current_income - current_expense
        self.worksheet.cell(row=row, column=6, value=current_balance)
        return current_balance
    
    def add_income(self, date, description, amount):
        """Menambahkan pemasukan"""
        day = self.get_day_name(date)
        next_row = self.worksheet.max_row + 1
        self.worksheet.cell(row=next_row, column=1, value=date)
        self.worksheet.cell(row=next_row, column=2, value=day)
        self.worksheet.cell(row=next_row, column=3, value=description)
        self.worksheet.cell(row=next_row, column=4, value=amount)
        self.worksheet.cell(row=next_row, column=5, value=0)
        balance = self.calculate_and_update_balance(next_row)
        self.workbook.save(self.excel_file)
        return balance
    
    def add_expense(self, date, description, amount):
        """Menambahkan pengeluaran"""
        day = self.get_day_name(date)
        next_row = self.worksheet.max_row + 1
        self.worksheet.cell(row=next_row, column=1, value=date)
        self.worksheet.cell(row=next_row, column=2, value=day)
        self.worksheet.cell(row=next_row, column=3, value=description)
        self.worksheet.cell(row=next_row, column=4, value=0)
        self.worksheet.cell(row=next_row, column=5, value=amount)
        balance = self.calculate_and_update_balance(next_row)
        self.workbook.save(self.excel_file)
        return balance
    
    def calculate_totals(self):
        """Menghitung total pemasukan, pengeluaran, dan saldo"""
        total_income = 0
        total_expense = 0
        current_balance = 0
        
        # Iterasi semua baris dengan data
        for row in range(2, self.worksheet.max_row + 1):
            income = self.worksheet.cell(row=row, column=4).value or 0
            expense = self.worksheet.cell(row=row, column=5).value or 0
            balance = self.worksheet.cell(row=row, column=6).value or 0
            
            total_income += income
            total_expense += expense
            current_balance = balance  # Saldo terakhir
        
        return total_income, total_expense, current_balance
    
    def build(self):
        self.title = 'Aplikasi Keuangan Katering'
        
        # Layout utama
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # Judul
        title = Label(text='Aplikasi Keuangan Katering', 
                     size_hint_y=None, 
                     height='50dp',
                     font_size='20sp',
                     bold=True)
        main_layout.add_widget(title)
        
        # Form input pemasukan
        income_layout = BoxLayout(orientation='vertical', size_hint_y=None, height='150dp')
        income_layout.add_widget(Label(text='Tambah Pemasukan', size_hint_y=None, height='30dp'))
        
        income_form = GridLayout(cols=2, spacing=10, size_hint_y=None, height='100dp')
        self.income_date = TextInput(hint_text='Tanggal (DD/MM/YYYY)')
        self.income_desc = TextInput(hint_text='Deskripsi')
        self.income_amount = TextInput(hint_text='Jumlah (Rp)')
        income_add_btn = Button(text='Tambah Pemasukan')
        income_add_btn.bind(on_press=self.add_income_entry)
        
        income_form.add_widget(Label(text='Tanggal:'))
        income_form.add_widget(self.income_date)
        income_form.add_widget(Label(text='Deskripsi:'))
        income_form.add_widget(self.income_desc)
        income_form.add_widget(Label(text='Jumlah:'))
        income_form.add_widget(self.income_amount)
        income_form.add_widget(Label(text=''))
        income_form.add_widget(income_add_btn)
        
        income_layout.add_widget(income_form)
        main_layout.add_widget(income_layout)
        
        # Form input pengeluaran
        expense_layout = BoxLayout(orientation='vertical', size_hint_y=None, height='150dp')
        expense_layout.add_widget(Label(text='Tambah Pengeluaran', size_hint_y=None, height='30dp'))
        
        expense_form = GridLayout(cols=2, spacing=10, size_hint_y=None, height='100dp')
        self.expense_date = TextInput(hint_text='Tanggal (DD/MM/YYYY)')
        self.expense_desc = TextInput(hint_text='Deskripsi')
        self.expense_amount = TextInput(hint_text='Jumlah (Rp)')
        expense_add_btn = Button(text='Tambah Pengeluaran')
        expense_add_btn.bind(on_press=self.add_expense_entry)
        
        expense_form.add_widget(Label(text='Tanggal:'))
        expense_form.add_widget(self.expense_date)
        expense_form.add_widget(Label(text='Deskripsi:'))
        expense_form.add_widget(self.expense_desc)
        expense_form.add_widget(Label(text='Jumlah:'))
        expense_form.add_widget(self.expense_amount)
        expense_form.add_widget(Label(text=''))
        expense_form.add_widget(expense_add_btn)
        
        expense_layout.add_widget(expense_form)
        main_layout.add_widget(expense_layout)
        
        # Tombol untuk menampilkan laporan
        report_btn = Button(text='Tampilkan Laporan Keuangan', size_hint_y=None, height='50dp')
        report_btn.bind(on_press=self.show_report)
        main_layout.add_widget(report_btn)
        
        # Tombol untuk menampilkan ringkasan
        summary_btn = Button(text='Tampilkan Ringkasan', size_hint_y=None, height='50dp')
        summary_btn.bind(on_press=self.show_summary)
        main_layout.add_widget(summary_btn)
        
        return main_layout
    
    def add_income_entry(self, instance):
        try:
            date = self.income_date.text
            description = self.income_desc.text
            amount = float(self.income_amount.text)
            
            if date and description and amount > 0:
                balance = self.add_income(date, description, amount)
                
                # Tampilkan popup konfirmasi
                popup = Popup(title='Pemasukan Ditambahkan',
                            content=Label(text=f'Pemasukan berhasil ditambahkan!\nSaldo terbaru: Rp {balance:,.0f}'),
                            size_hint=(0.8, 0.4))
                popup.open()
                
                # Kosongkan field
                self.income_date.text = ''
                self.income_desc.text = ''
                self.income_amount.text = ''
            else:
                popup = Popup(title='Error',
                            content=Label(text='Harap isi semua field dengan benar!'),
                            size_hint=(0.8, 0.4))
                popup.open()
        except ValueError:
            popup = Popup(title='Error',
                        content=Label(text='Jumlah harus berupa angka!'),
                        size_hint=(0.8, 0.4))
            popup.open()
    
    def add_expense_entry(self, instance):
        try:
            date = self.expense_date.text
            description = self.expense_desc.text
            amount = float(self.expense_amount.text)
            
            if date and description and amount > 0:
                balance = self.add_expense(date, description, amount)
                
                # Tampilkan popup konfirmasi
                popup = Popup(title='Pengeluaran Ditambahkan',
                            content=Label(text=f'Pengeluaran berhasil ditambahkan!\nSaldo terbaru: Rp {balance:,.0f}'),
                            size_hint=(0.8, 0.4))
                popup.open()
                
                # Kosongkan field
                self.expense_date.text = ''
                self.expense_desc.text = ''
                self.expense_amount.text = ''
            else:
                popup = Popup(title='Error',
                            content=Label(text='Harap isi semua field dengan benar!'),
                            size_hint=(0.8, 0.4))
                popup.open()
        except ValueError:
            popup = Popup(title='Error',
                        content=Label(text='Jumlah harus berupa angka!'),
                        size_hint=(0.8, 0.4))
            popup.open()
    
    def show_report(self, instance):
        # Buat layout untuk laporan
        report_layout = BoxLayout(orientation='vertical', padding=10)
        
        # Header laporan
        header = Label(text='LAPORAN KEUANGAN KATERING', 
                      size_hint_y=None, 
                      height='40dp',
                      font_size='18sp',
                      bold=True)
        report_layout.add_widget(header)
        
        # Scroll view untuk data
        scroll = ScrollView()
        grid = GridLayout(cols=6, spacing=5, size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))
        
        # Header tabel
        grid.add_widget(Label(text='Tanggal', bold=True))
        grid.add_widget(Label(text='Hari', bold=True))
        grid.add_widget(Label(text='Deskripsi', bold=True))
        grid.add_widget(Label(text='Pemasukan', bold=True))
        grid.add_widget(Label(text='Pengeluaran', bold=True))
        grid.add_widget(Label(text='Saldo', bold=True))
        
        # Data transaksi
        for row in range(2, self.worksheet.max_row + 1):
            date = self.worksheet.cell(row=row, column=1).value or ""
            day = self.worksheet.cell(row=row, column=2).value or ""
            desc = self.worksheet.cell(row=row, column=3).value or ""
            income = self.worksheet.cell(row=row, column=4).value or 0
            expense = self.worksheet.cell(row=row, column=5).value or 0
            balance = self.worksheet.cell(row=row, column=6).value or 0
            
            grid.add_widget(Label(text=date))
            grid.add_widget(Label(text=day))
            grid.add_widget(Label(text=desc))
            grid.add_widget(Label(text=f"Rp {income:,.0f}"))
            grid.add_widget(Label(text=f"Rp {expense:,.0f}"))
            grid.add_widget(Label(text=f"Rp {balance:,.0f}"))
        
        scroll.add_widget(grid)
        report_layout.add_widget(scroll)
        
        # Tombol tutup
        close_btn = Button(text='Tutup', size_hint_y=None, height='50dp')
        report_layout.add_widget(close_btn)
        
        # Buat popup
        popup = Popup(title='Laporan Keuangan',
                     content=report_layout,
                     size_hint=(0.9, 0.9))
        close_btn.bind(on_press=popup.dismiss)
        popup.open()
    
    def show_summary(self, instance):
        total_income, total_expense, current_balance = self.calculate_totals()
        
        summary_text = f"Ringkasan Keuangan:\n\n"
        summary_text += f"Total Pemasukan: Rp {total_income:,.0f}\n"
        summary_text += f"Total Pengeluaran: Rp {total_expense:,.0f}\n"
        summary_text += f"Sisa Uang: Rp {current_balance:,.0f}"
        
        popup = Popup(title='Ringkasan Keuangan',
                     content=Label(text=summary_text),
                     size_hint=(0.8, 0.6))
        popup.open()

if __name__ == '__main__':
    CateringFinanceApp().run()
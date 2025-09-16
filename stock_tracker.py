import openpyxl
from datetime import datetime

class StockCardTracker:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.workbook = openpyxl.load_workbook(excel_file)
        self.worksheet = self.workbook.active
        
    def display_stock_card(self):
        """Display the current stock card data"""
        print("Current Stock Card:")
        print("-" * 80)
        
        # Print header
        header = []
        for col in range(1, 6):  # Columns A to E
            cell_value = self.worksheet.cell(row=1, column=col).value
            header.append(cell_value if cell_value else "")
        print(f"{header[0]:<12} {header[1]:<30} {header[2]:<10} {header[3]:<10} {header[4]:<10}")
        print("-" * 80)
        
        # Print data rows
        for row in range(2, self.worksheet.max_row + 1):
            row_data = []
            for col in range(1, 6):  # Columns A to E
                cell_value = self.worksheet.cell(row=row, column=col).value
                row_data.append(cell_value if cell_value else "")
            
            # Format the output based on column types
            if str(row_data[0]).isdigit():  # If first column is a number
                print(f"{row_data[0]:<12} {row_data[1]:<30} {row_data[2]:<10} {row_data[3]:<10} {row_data[4]:<10}")
            else:
                print(f"{row_data[0]:<12} {row_data[1]:<30} {row_data[2]:<10} {row_data[3]:<10} {row_data[4]:<10}")
    
    def get_current_balance(self):
        """Get the current stock balance from the last row"""
        last_row = self.worksheet.max_row
        balance_cell = self.worksheet.cell(row=last_row, column=5)  # Column E (Sisa)
        return balance_cell.value if balance_cell.value else 0
    
    def add_stock_entry(self, date, description, incoming, outgoing):
        """Add a new stock entry"""
        # Find the next available row
        next_row = self.worksheet.max_row + 1
        
        # Add the entry number (based on previous row)
        prev_row_num = self.worksheet.cell(row=next_row-1, column=1).value
        entry_num = int(prev_row_num) + 1 if isinstance(prev_row_num, (int, float)) else next_row - 2
        
        # Write data to the worksheet
        self.worksheet.cell(row=next_row, column=1, value=entry_num)
        self.worksheet.cell(row=next_row, column=2, value=date)
        self.worksheet.cell(row=next_row, column=3, value=description)
        self.worksheet.cell(row=next_row, column=4, value=incoming)
        self.worksheet.cell(row=next_row, column=5, value=outgoing)
        
        # Calculate the new balance
        current_balance = self.get_current_balance()
        new_balance = current_balance + incoming - outgoing
        self.worksheet.cell(row=next_row, column=6, value=new_balance)  # Column F for balance
        
        print(f"Added new entry #{entry_num}: {description} on {date}")
        print(f"New balance: {new_balance}")
    
    def save_workbook(self):
        """Save the workbook to the file"""
        self.workbook.save(self.excel_file)
        print("Stock card saved successfully!")

def main():
    # Initialize the stock card tracker
    tracker = StockCardTracker("kartu_stok.xlsx")
    
    # Display current stock card
    tracker.display_stock_card()
    
    # Show current balance
    balance = tracker.get_current_balance()
    print(f"\nCurrent Balance: {balance}")
    
    # Menu for user interaction
    while True:
        print("\nStock Card Tracker Menu:")
        print("1. Add new stock entry")
        print("2. Display stock card")
        print("3. Show current balance")
        print("4. Save and exit")
        print("5. Exit without saving")
        
        choice = input("Enter your choice (1-5): ")
        
        if choice == "1":
            date = input("Enter date (DD/MM/YYYY): ")
            description = input("Enter description: ")
            try:
                incoming = float(input("Enter incoming quantity (0 if none): "))
                outgoing = float(input("Enter outgoing quantity (0 if none): "))
                tracker.add_stock_entry(date, description, incoming, outgoing)
            except ValueError:
                print("Invalid input. Please enter numeric values for quantities.")
        
        elif choice == "2":
            tracker.display_stock_card()
        
        elif choice == "3":
            balance = tracker.get_current_balance()
            print(f"Current Balance: {balance}")
        
        elif choice == "4":
            tracker.save_workbook()
            print("Goodbye!")
            break
        
        elif choice == "5":
            print("Exiting without saving changes.")
            break
        
        else:
            print("Invalid choice. Please enter a number between 1 and 5.")

if __name__ == "__main__":
    main()
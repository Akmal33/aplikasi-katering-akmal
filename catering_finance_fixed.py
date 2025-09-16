import openpyxl
from datetime import datetime, timedelta
import os

class CateringFinanceTracker:
    def __init__(self, excel_file="catering_finance_fixed.xlsx"):
        self.excel_file = excel_file
        self.workbook = None
        self.worksheet = None
        self.initialize_workbook()
    
    def initialize_workbook(self):
        """Inisialisasi workbook dan worksheet"""
        if os.path.exists(self.excel_file):
            self.workbook = openpyxl.load_workbook(self.excel_file)
            if "Keuangan_Katering" in self.workbook.sheetnames:
                self.worksheet = self.workbook["Keuangan_Katering"]
            else:
                self.worksheet = self.workbook.create_sheet("Keuangan_Katering")
        else:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Keuangan_Katering"
            self.create_headers()
    
    def create_headers(self):
        """Membuat header untuk worksheet"""
        headers = ["Tanggal", "Hari", "Deskripsi", "Pemasukan (Rp)", "Pengeluaran (Rp)", "Saldo (Rp)"]
        for col, header in enumerate(headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)
    
    def add_income(self, date, day, description, amount):
        """Menambahkan pemasukan"""
        next_row = self.worksheet.max_row + 1
        self.worksheet.cell(row=next_row, column=1, value=date)
        self.worksheet.cell(row=next_row, column=2, value=day)
        self.worksheet.cell(row=next_row, column=3, value=description)
        self.worksheet.cell(row=next_row, column=4, value=amount)
        self.worksheet.cell(row=next_row, column=5, value=0)
        self.calculate_and_update_balance(next_row)
        print(f"Pemasukan ditambahkan: {description} - Rp {amount:,}")
    
    def add_expense(self, date, day, description, amount):
        """Menambahkan pengeluaran"""
        next_row = self.worksheet.max_row + 1
        self.worksheet.cell(row=next_row, column=1, value=date)
        self.worksheet.cell(row=next_row, column=2, value=day)
        self.worksheet.cell(row=next_row, column=3, value=description)
        self.worksheet.cell(row=next_row, column=4, value=0)
        self.worksheet.cell(row=next_row, column=5, value=amount)
        self.calculate_and_update_balance(next_row)
        print(f"Pengeluaran ditambahkan: {description} - Rp {amount:,}")
    
    def calculate_and_update_balance(self, row):
        """Menghitung dan memperbarui saldo"""
        current_income = self.worksheet.cell(row=row, column=4).value or 0
        current_expense = self.worksheet.cell(row=row, column=5).value or 0
        
        if row == 2:  # Baris pertama dengan data
            previous_balance = 0
        else:
            previous_balance = self.worksheet.cell(row=row-1, column=6).value or 0
        
        current_balance = previous_balance + current_income - current_expense
        self.worksheet.cell(row=row, column=6, value=current_balance)
    
    def calculate_totals(self):
        """Menghitung total pemasukan, pengeluaran, dan saldo"""
        total_income = 0
        total_expense = 0
        current_balance = 0
        
        # Iterasi semua baris dengan data
        for row in range(2, self.worksheet.max_row + 1):
            income = self.worksheet.cell(row=row, column=4).value or 0
            expense = self.worksheet.cell(row=row, column=5).value or 0
            balance = self.worksheet.cell(row=row, column=6).value or 0
            
            total_income += income
            total_expense += expense
            current_balance = balance  # Saldo terakhir
        
        return total_income, total_expense, current_balance
    
    def display_weekly_report(self):
        """Menampilkan laporan mingguan"""
        print("\n" + "="*70)
        print("LAPORAN KEUANGAN KATERING MINGGUAN")
        print("="*70)
        
        # Header tabel
        print(f"{'Tanggal':<12} {'Hari':<10} {'Deskripsi':<25} {'Pemasukan':<12} {'Pengeluaran':<12} {'Saldo':<12}")
        print("-"*70)
        
        # Data transaksi
        for row in range(2, self.worksheet.max_row + 1):
            date = self.worksheet.cell(row=row, column=1).value or ""
            day = self.worksheet.cell(row=row, column=2).value or ""
            desc = self.worksheet.cell(row=row, column=3).value or ""
            income = self.worksheet.cell(row=row, column=4).value or 0
            expense = self.worksheet.cell(row=row, column=5).value or 0
            balance = self.worksheet.cell(row=row, column=6).value or 0
            
            print(f"{date:<12} {day:<10} {desc:<25} {income:>10,.0f}   {expense:>10,.0f}   {balance:>10,.0f}")
        
        # Total
        total_income, total_expense, current_balance = self.calculate_totals()
        print("-"*70)
        print(f"{'TOTAL':<49} {total_income:>10,.0f}   {total_expense:>10,.0f}   {current_balance:>10,.0f}")
        print("="*70)
    
    def save_workbook(self):
        """Menyimpan workbook ke file"""
        try:
            self.workbook.save(self.excel_file)
            print("Data berhasil disimpan ke file Excel!")
            return True
        except Exception as e:
            print(f"Error saat menyimpan file: {e}")
            return False

def get_day_name(date_str):
    """Mendapatkan nama hari dari tanggal"""
    days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    try:
        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
        return days[date_obj.weekday()]
    except ValueError:
        return "Tidak valid"

def fixed_example():
    """Contoh penggunaan aplikasi yang diperbaiki"""
    print("=== CONTOH PENGGUNAAN APLIKASI KEUANGAN KATERING (Diperbaiki) ===\n")
    
    # Inisialisasi tracker
    tracker = CateringFinanceTracker()
    
    # Simulasi input pengguna (minggu pertama)
    print("Minggu Pertama:")
    print("1. Menambahkan pemasukan awal:")
    tracker.add_income("01/09/2025", "Senin", "Pembayaran pesanan catering", 2500000)
    
    print("\n2. Menambahkan pengeluaran harian:")
    tracker.add_expense("01/09/2025", "Senin", "Belanja bahan baku", 450000)
    tracker.add_expense("01/09/2025", "Senin", "Biaya transportasi", 75000)
    
    print("\n3. Menambahkan pemasukan tambahan:")
    tracker.add_income("02/09/2025", "Selasa", "Pembayaran catering harian", 1800000)
    
    print("\n4. Menambahkan pengeluaran lainnya:")
    tracker.add_expense("02/09/2025", "Selasa", "Belanja bahan baku", 380000)
    tracker.add_expense("02/09/2025", "Selasa", "Biaya gas dan listrik", 90000)
    
    print("\n5. Menambahkan pemasukan akhir minggu:")
    tracker.add_income("03/09/2025", "Rabu", "Pembayaran catering pernikahan", 4200000)
    
    print("\n6. Menambahkan pengeluaran akhir minggu:")
    tracker.add_expense("03/09/2025", "Rabu", "Belanja bahan baku", 520000)
    tracker.add_expense("03/09/2025", "Rabu", "Upah karyawan", 950000)
    
    # Tampilkan laporan
    tracker.display_weekly_report()
    
    # Hitung total
    total_income, total_expense, current_balance = tracker.calculate_totals()
    print(f"\nRingkasan Keuangan:")
    print(f"Total Pemasukan: Rp {total_income:,}")
    print(f"Total Pengeluaran: Rp {total_expense:,}")
    print(f"Sisa Uang: Rp {current_balance:,}")
    
    # Simpan data
    tracker.save_workbook()
    
    print("\n=== PENJELASAN PERHITUNGAN ===")
    print("Perhitungan Saldo:")
    print("1. Saldo awal: Rp 0")
    print("2. + Pemasukan (2.500.000) = Rp 2.500.000")
    print("3. - Pengeluaran (450.000 + 75.000) = Rp 1.975.000")
    print("4. + Pemasukan (1.800.000) = Rp 3.775.000")
    print("5. - Pengeluaran (380.000 + 90.000) = Rp 3.305.000")
    print("6. + Pemasukan (4.200.000) = Rp 7.505.000")
    print("7. - Pengeluaran (520.000 + 950.000) = Rp 6.035.000")
    print("\nSisa Uang Akhir: Rp 6.035.000")

if __name__ == "__main__":
    fixed_example()
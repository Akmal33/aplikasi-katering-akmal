import openpyxl
from datetime import datetime
import os

class StockCardTracker:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        if os.path.exists(excel_file):
            self.workbook = openpyxl.load_workbook(excel_file)
        else:
            # Create a new workbook if file doesn't exist
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Kartu Stok"
            # Add headers
            headers = ["No.", "Tanggal", "Keterangan", "Masuk", "Keluar", "Sisa"]
            for col, header in enumerate(headers, 1):
                self.worksheet.cell(row=1, column=col, value=header)
        
        self.worksheet = self.workbook["Kartu Stok"] if "Kartu Stok" in self.workbook.sheetnames else self.workbook.active
        
    def display_stock_card(self):
        """Display the current stock card data"""
        print("Current Stock Card:")
        print("-" * 80)
        
        # Check if there's data
        if self.worksheet.max_row == 1:
            print("No data available.")
            return
            
        # Print header
        header = []
        for col in range(1, 7):  # Columns A to F
            cell_value = self.worksheet.cell(row=1, column=col).value
            header.append(str(cell_value) if cell_value is not None else "")
        print(f"{header[0]:<5} {header[1]:<12} {header[2]:<25} {header[3]:<10} {header[4]:<10} {header[5]:<10}")
        print("-" * 80)
        
        # Print data rows
        for row in range(2, self.worksheet.max_row + 1):
            row_data = []
            for col in range(1, 7):  # Columns A to F
                cell_value = self.worksheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            
            print(f"{row_data[0]:<5} {row_data[1]:<12} {row_data[2]:<25} {row_data[3]:<10} {row_data[4]:<10} {row_data[5]:<10}")
    
    def get_current_balance(self):
        """Get the current stock balance from the last row"""
        if self.worksheet.max_row <= 1:
            return 0
            
        last_row = self.worksheet.max_row
        balance_cell = self.worksheet.cell(row=last_row, column=6)  # Column F (Sisa)
        try:
            return float(balance_cell.value) if balance_cell.value is not None else 0
        except (ValueError, TypeError):
            return 0
    
    def add_stock_entry(self, date, description, incoming, outgoing):
        """Add a new stock entry"""
        # Find the next available row
        next_row = self.worksheet.max_row + 1
        
        # Add the entry number (based on previous row)
        if next_row == 2:  # First entry
            entry_num = 1
        else:
            prev_row_num = self.worksheet.cell(row=next_row-1, column=1).value
            try:
                entry_num = int(prev_row_num) + 1
            except (ValueError, TypeError):
                entry_num = next_row - 1
        
        # Write data to the worksheet
        self.worksheet.cell(row=next_row, column=1, value=entry_num)
        self.worksheet.cell(row=next_row, column=2, value=date)
        self.worksheet.cell(row=next_row, column=3, value=description)
        self.worksheet.cell(row=next_row, column=4, value=float(incoming))
        self.worksheet.cell(row=next_row, column=5, value=float(outgoing))
        
        # Calculate the new balance
        current_balance = self.get_current_balance()
        new_balance = current_balance + float(incoming) - float(outgoing)
        self.worksheet.cell(row=next_row, column=6, value=new_balance)  # Column F for balance
        
        print(f"Added new entry #{entry_num}: {description} on {date}")
        print(f"New balance: {new_balance}")
        return entry_num
    
    def save_workbook(self):
        """Save the workbook to the file"""
        try:
            self.workbook.save(self.excel_file)
            print("Stock card saved successfully!")
            return True
        except Exception as e:
            print(f"Error saving file: {e}")
            return False

def main():
    # Initialize the stock card tracker with the proper Excel file
    tracker = StockCardTracker("kartu_stok_proper.xlsx")
    
    # Display current stock card
    tracker.display_stock_card()
    
    # Show current balance
    balance = tracker.get_current_balance()
    print(f"\nCurrent Balance: {balance}")
    
    # Menu for user interaction
    while True:
        print("\nStock Card Tracker Menu:")
        print("1. Add new stock entry")
        print("2. Display stock card")
        print("3. Show current balance")
        print("4. Save and exit")
        print("5. Exit without saving")
        
        choice = input("Enter your choice (1-5): ")
        
        if choice == "1":
            date = input("Enter date (DD/MM/YYYY): ")
            description = input("Enter description: ")
            try:
                incoming = float(input("Enter incoming quantity (0 if none): ") or "0")
                outgoing = float(input("Enter outgoing quantity (0 if none): ") or "0")
                tracker.add_stock_entry(date, description, incoming, outgoing)
            except ValueError:
                print("Invalid input. Please enter numeric values for quantities.")
        
        elif choice == "2":
            tracker.display_stock_card()
        
        elif choice == "3":
            balance = tracker.get_current_balance()
            print(f"Current Balance: {balance}")
        
        elif choice == "4":
            if tracker.save_workbook():
                print("Goodbye!")
                break
            else:
                print("Failed to save. Do you want to exit without saving? (y/n)")
                if input().lower() == 'y':
                    break
        
        elif choice == "5":
            print("Exiting without saving changes.")
            break
        
        else:
            print("Invalid choice. Please enter a number between 1 and 5.")

if __name__ == "__main__":
    main()
import openpyxl
import os
from datetime import datetime

class ConsoleCateringFinanceApp:
    def __init__(self):
        self.excel_file = "catering_finance_console.xlsx"
        # Hapus file sebelumnya jika ada
        if os.path.exists(self.excel_file):
            os.remove(self.excel_file)
        self.initialize_workbook()
    
    def initialize_workbook(self):
        """Inisialisasi workbook dan worksheet"""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Keuangan_Katering"
        self.create_headers()
    
    def create_headers(self):
        """Membuat header untuk worksheet"""
        headers = ["Tanggal", "Hari", "Deskripsi", "Pemasukan (Rp)", "Pengeluaran (Rp)", "Saldo (Rp)"]
        for col, header in enumerate(headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)
    
    def get_day_name(self, date_str):
        """Mendapatkan nama hari dari tanggal"""
        days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        try:
            date_obj = datetime.strptime(date_str, "%d/%m/%Y")
            return days[date_obj.weekday()]
        except ValueError:
            return "Tidak valid"
    
    def add_income(self, date, description, amount):
        """Menambahkan pemasukan"""
        day = self.get_day_name(date)
        next_row = self.worksheet.max_row + 1
        self.worksheet.cell(row=next_row, column=1, value=date)
        self.worksheet.cell(row=next_row, column=2, value=day)
        self.worksheet.cell(row=next_row, column=3, value=description)
        self.worksheet.cell(row=next_row, column=4, value=amount)
        self.worksheet.cell(row=next_row, column=5, value=0)
        
        # Hitung saldo baru
        if next_row == 2:  # Baris pertama dengan data
            previous_balance = 0
        else:
            previous_balance = self.worksheet.cell(row=next_row-1, column=6).value or 0
        
        current_balance = previous_balance + amount
        self.worksheet.cell(row=next_row, column=6, value=current_balance)
        
        self.workbook.save(self.excel_file)
        return current_balance
    
    def add_expense(self, date, description, amount):
        """Menambahkan pengeluaran"""
        day = self.get_day_name(date)
        next_row = self.worksheet.max_row + 1
        self.worksheet.cell(row=next_row, column=1, value=date)
        self.worksheet.cell(row=next_row, column=2, value=day)
        self.worksheet.cell(row=next_row, column=3, value=description)
        self.worksheet.cell(row=next_row, column=4, value=0)
        self.worksheet.cell(row=next_row, column=5, value=amount)
        
        # Hitung saldo baru
        if next_row == 2:  # Baris pertama dengan data
            previous_balance = 0
        else:
            previous_balance = self.worksheet.cell(row=next_row-1, column=6).value or 0
        
        current_balance = previous_balance - amount
        self.worksheet.cell(row=next_row, column=6, value=current_balance)
        
        self.workbook.save(self.excel_file)
        return current_balance
    
    def calculate_totals(self):
        """Menghitung total pemasukan, pengeluaran, dan saldo"""
        total_income = 0
        total_expense = 0
        current_balance = 0
        
        # Iterasi semua baris dengan data
        for row in range(2, self.worksheet.max_row + 1):
            income = self.worksheet.cell(row=row, column=4).value or 0
            expense = self.worksheet.cell(row=row, column=5).value or 0
            balance = self.worksheet.cell(row=row, column=6).value or 0
            
            total_income += income
            total_expense += expense
            current_balance = balance  # Saldo terakhir
        
        return total_income, total_expense, current_balance
    
    def display_report(self):
        """Menampilkan laporan keuangan"""
        print("\n" + "="*70)
        print("LAPORAN KEUANGAN KATERING")
        print("="*70)
        
        # Header tabel
        print(f"{'Tanggal':<12} {'Hari':<10} {'Deskripsi':<25} {'Pemasukan':<12} {'Pengeluaran':<12} {'Saldo':<12}")
        print("-"*70)
        
        # Data transaksi
        for row in range(2, self.worksheet.max_row + 1):
            date = self.worksheet.cell(row=row, column=1).value or ""
            day = self.worksheet.cell(row=row, column=2).value or ""
            desc = self.worksheet.cell(row=row, column=3).value or ""
            income = self.worksheet.cell(row=row, column=4).value or 0
            expense = self.worksheet.cell(row=row, column=5).value or 0
            balance = self.worksheet.cell(row=row, column=6).value or 0
            
            print(f"{date:<12} {day:<10} {desc:<25} {income:>10,.0f}   {expense:>10,.0f}   {balance:>10,.0f}")
        
        # Total
        total_income, total_expense, current_balance = self.calculate_totals()
        print("-"*70)
        print(f"{'TOTAL':<49} {total_income:>10,.0f}   {total_expense:>10,.0f}   {current_balance:>10,.0f}")
        print("="*70)

def demo_console_app():
    """Demo aplikasi console"""
    print("=== DEMO APLIKASI KEUANGAN KATERING (Versi Console) ===")
    print("Ini menunjukkan cara kerja aplikasi mobile yang akan dibangun menjadi .apk")
    print()
    
    # Inisialisasi aplikasi
    app = ConsoleCateringFinanceApp()
    
    # Tambahkan beberapa transaksi contoh
    print("Menambahkan transaksi contoh:")
    
    # Pemasukan
    balance = app.add_income("01/09/2025", "Pembayaran pesanan catering", 2500000)
    print(f"[OK] Pemasukan: Pembayaran pesanan catering - Rp 2,500,000 (Saldo: Rp {balance:,.0f})")
    
    # Pengeluaran
    balance = app.add_expense("01/09/2025", "Belanja bahan baku", 450000)
    print(f"[OK] Pengeluaran: Belanja bahan baku - Rp 450,000 (Saldo: Rp {balance:,.0f})")
    
    balance = app.add_expense("01/09/2025", "Biaya transportasi", 75000)
    print(f"[OK] Pengeluaran: Biaya transportasi - Rp 75,000 (Saldo: Rp {balance:,.0f})")
    
    # Pemasukan lainnya
    balance = app.add_income("02/09/2025", "Pembayaran catering harian", 1800000)
    print(f"[OK] Pemasukan: Pembayaran catering harian - Rp 1,800,000 (Saldo: Rp {balance:,.0f})")
    
    # Pengeluaran lainnya
    balance = app.add_expense("02/09/2025", "Belanja bahan baku", 380000)
    print(f"[OK] Pengeluaran: Belanja bahan baku - Rp 380,000 (Saldo: Rp {balance:,.0f})")
    
    balance = app.add_expense("02/09/2025", "Biaya gas dan listrik", 90000)
    print(f"[OK] Pengeluaran: Biaya gas dan listrik - Rp 90,000 (Saldo: Rp {balance:,.0f})")
    
    # Pemasukan besar
    balance = app.add_income("03/09/2025", "Pembayaran catering pernikahan", 4200000)
    print(f"[OK] Pemasukan: Pembayaran catering pernikahan - Rp 4,200,000 (Saldo: Rp {balance:,.0f})")
    
    # Pengeluaran besar
    balance = app.add_expense("03/09/2025", "Belanja bahan baku", 520000)
    print(f"[OK] Pengeluaran: Belanja bahan baku - Rp 520,000 (Saldo: Rp {balance:,.0f})")
    
    balance = app.add_expense("03/09/2025", "Upah karyawan", 950000)
    print(f"[OK] Pengeluaran: Upah karyawan - Rp 950,000 (Saldo: Rp {balance:,.0f})")
    
    # Tampilkan laporan
    app.display_report()
    
    # Tampilkan ringkasan
    total_income, total_expense, current_balance = app.calculate_totals()
    print(f"\nRingkasan Keuangan:")
    print(f"Total Pemasukan: Rp {total_income:,.0f}")
    print(f"Total Pengeluaran: Rp {total_expense:,.0f}")
    print(f"Sisa Uang: Rp {current_balance:,.0f}")
    
    print(f"\nData telah disimpan ke file: {app.excel_file}")
    
    print("\n" + "="*50)
    print("INFORMASI PENTING:")
    print("="*50)
    print("Versi mobile (.apk) memiliki antarmuka grafis dengan:")
    print("1. Form input untuk pemasukan")
    print("2. Form input untuk pengeluaran")
    print("3. Tombol untuk melihat laporan")
    print("4. Tombol untuk melihat ringkasan")
    print("5. Penyimpanan data otomatis ke Excel")
    print("\nUntuk membangun .apk, ikuti instruksi di README_MOBILE.md")
    print("="*50)

if __name__ == "__main__":
    demo_console_app()